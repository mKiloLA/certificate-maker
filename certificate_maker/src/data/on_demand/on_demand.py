import os

import pandas as pd
from typing import Any
import openpyxl

from certificate_maker.src.data.on_demand.report_entry import ReportEntry
from certificate_maker.src.data.ref import us_state_to_abbrev
from certificate_maker.src.exception_types import IncorrectDateTimeFormat, MissingSubmissionData, MissingEvaluationData, MalformedCROString, MalformedEvaluationQuestionResponse, ReferenceFileMissingSheet, ReferenceFileMissingCourse


def _abbreviate_rating(rating: Any) -> str:
    """
    Convert a rating string to abbreviated form using first letter of each word.
    Example: "Exceeded Expectation" -> "EE", "Met Expectation" -> "ME"
    """
    if pd.isna(rating) or not rating:
        return ""
    
    rating_str = str(rating).strip()
    words = rating_str.split()
    abbreviated = "".join(word[0].upper() for word in words if word)
    return abbreviated


def _parse_evaluation_info(row: pd.Series, rating_columns: list[str]) -> str | None:
    """
    Extract course evaluation from row by abbreviating all "Please rate..." responses.
    Abbreviates each response by taking the first letter of each word.
    Example: "Exceeded Expectation" -> "EE", "Met Expectation" -> "ME"
    Returns comma-separated abbreviations, or None if no data.
    """
    try:
        abbreviations: list[str] = [
            _abbreviate_rating(row.get(col)) for col in rating_columns
        ]
        # Filter out empty strings and join with ", "
        result = ", ".join(abbr for abbr in abbreviations if abbr)
        return result if result else None
    except Exception as e:
        raise MalformedEvaluationQuestionResponse(e)

def _parse_cro_info(cro_string: Any) -> list[tuple[str, str]]:
    """
    Extract all state and bar number pairs from CRO string.
    Format: "State: BarNumber - Hours" or multiple entries separated by ", "
    Example: "Georgia: 289725 - 2 hours" -> [("GA", "289725")]
    Example: "Tennessee: 017801 - 1 hours, Tennessee: 017801 - 2 hours" -> [("TN", "017801")]
    Returns list of tuples, deduped by state.
    """
    if pd.isna(cro_string) or not cro_string:
        return []
    
    try:
        cro_str = str(cro_string).strip()
        entries: list[tuple[str, str]] = []
        seen_states: set[str] = set()
        
        # Split by comma and space to get individual entries
        for entry_text in cro_str.split(", "):
            entry_text = entry_text.strip()
            if not entry_text:
                continue
            
            # Split by colon to get state name and rest
            parts = entry_text.split(":")
            if len(parts) < 2:
                continue
            
            state_name = parts[0].strip()
            # Find the state abbreviation
            state_abbrev = us_state_to_abbrev.get(state_name)
            
            # Skip if we've already seen this state
            if state_abbrev in seen_states:
                continue
            
            # Extract bar number from the second part
            bar_part = parts[1].split("-")[0].strip()
            
            if state_abbrev and bar_part:
                entries.append((state_abbrev, bar_part))
                seen_states.add(state_abbrev)
        
        return entries
    except Exception:
        raise MalformedCROString(cro_string)


def _export_to_excel(report_entries: list[ReportEntry]) -> None:
    """Export report entries to an Excel file with specified headers."""
    data = []
    
    for entry in report_entries:
        data.append({
            "Last Name": entry.last_name,
            "First Name": entry.first_name,
            "Email": entry.email,
            "Bar State": entry.state,
            "Bar Number": entry.bar_number,
            "Course ID": entry.course_id,
            "Course Title": entry.course_title,
            "Hours": entry.course_hours,
            "Course Completed": entry.course_completed_date,
            "Course Evaluation": entry.course_evaluation,
            "Submitted to State": "",
            "Attendance Pd": "",
            "Submitted by": "",
            "Notes": ""
        })
    
    df = pd.DataFrame(data)
    
    # Write to Excel file with today's date
    from datetime import datetime
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    current_date = datetime.now().strftime("%m%d%Y")
    path_to_form = os.path.join(os.path.expanduser('~'), f"OnDemand/Output/on_demand_report_{current_date}.xlsx")
    os.makedirs(os.path.dirname(path_to_form), exist_ok=True)
    
    # Write dataframe to Excel normally (headers at row 1)
    df.to_excel(path_to_form, index=False, engine="openpyxl", sheet_name="On Demand Report")
    
    # Load the workbook and format it
    wb = openpyxl.load_workbook(path_to_form)
    ws = wb.active

    if (ws is None):
        raise ReferenceFileMissingSheet("On Demand Report")
    
    # Insert 2 rows at the top for title and date range (this pushes headers to row 3)
    ws.insert_rows(1, 2)
    
    # Title row
    ws["A1"] = "Comedian of Law, On Demand Hours Completed"
    
    # Date range row
    ws["A2"] = "<date range>"
    
    # Format title and date range
    arial_font_bold_underline = Font(name="Arial", size=12, bold=True, underline="single")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Merge cells for title (across all columns)
    max_col = get_column_letter(len(df.columns))
    ws.merge_cells(f"A1:{max_col}1")
    ws["A1"].font = arial_font_bold_underline
    ws["A1"].alignment = center_alignment
    
    # Merge cells for date range
    ws.merge_cells(f"A2:{max_col}2")
    ws["A2"].font = arial_font_bold_underline
    ws["A2"].alignment = center_alignment
    
    # Format table headers (bold, Arial, size 12) - now at row 3
    arial_font_bold = Font(name="Arial", size=12, bold=True)
    arial_font = Font(name="Arial", size=12)
    
    for cell in ws[3]:
        cell.font = arial_font_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format all data cells to Arial, size 12
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = arial_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Remove any existing tables first
    if hasattr(ws, '_tables'):
        ws._tables.clear()
    
    # Create table format starting from row 3 (headers) to last row with data
    tab = Table(displayName="OnDemandReport", ref=f"A3:{max_col}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Highlight rows with expired courses in red
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for idx, entry in enumerate(report_entries):
        if entry.course_expired:
            # Excel row is data_row_index + 3 (1 for header, 2 for title/date)
            excel_row = idx + 4
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = red_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        if column[0].column is None:
            continue
        column_letter = get_column_letter(column[0].column)
        column_header = ws[f"{column_letter}3"].value
        
        for cell in column[2:]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3
    
    wb.save(path_to_form)
    print(f"Excel report exported to {path_to_form}")


def create_on_demand_report(
    attendance_file: str, evaluation_file: str, reference_file: str
) -> None:
    """Creates a list of ReportEntry objects from the on-demand report."""
    # Read attendance file
    attendance_df: pd.DataFrame = pd.read_csv(attendance_file)
    
    report_entries: list[ReportEntry] = []
    
    for _, row in attendance_df.iterrows():
        # Extract state and bar number from CRO string
        cro_entries: list[tuple[str, str]] = _parse_cro_info(row.get("Submitted CRO (Member Number) and Hours"))

        # Split course title on last colon and take everything before it
        course_title: Any = row.get("Chapter Title")
        if course_title and ':' in course_title:
            course_title = course_title.rsplit(':', 1)[0]

        date_submitted_string = (row.get("Date Submitted"))
        course_completed_datetime: Any = None
        try:
            if (date_submitted_string is not None):
                course_completed_datetime = pd.to_datetime(date_submitted_string, format="%d-%b-%y")
        except Exception:
            raise IncorrectDateTimeFormat(date_submitted_string)
        
        # Iterate through each state and bar number pair to create a ReportEntry for each
        for state, bar_number in cro_entries:
            entry: ReportEntry = ReportEntry(
                first_name=row.get("First Name"),
                last_name=row.get("Last Name"),
                email=row.get("Email"),
                state=state,
                bar_number=bar_number,
                course_title=course_title,
                course_completed_date=course_completed_datetime
            )
            report_entries.append(entry)
    
    del attendance_df  # Free memory

    # Read evaluation file
    evaluation_df: pd.DataFrame = pd.read_csv(evaluation_file)
    
    # Get all columns that start with "Please rate" plus the recommendation column
    rating_columns: list[str] = [col for col in evaluation_df.columns if col.startswith("Please rate")]
    if "Would you recommend Comedian of Law to others?" in evaluation_df.columns:
        rating_columns.append("Would you recommend Comedian of Law to others?")

    # Compare email and course title to find the evaluation for each report entry
    # There may be multiple entries with the same email and course title and we want to update all of them that match
    for _, row in evaluation_df.iterrows():
        email = row.get("Email")
        course_title = row.get("Course Title")
        course_evaluation = _parse_evaluation_info(row, rating_columns)

        # Track if we found a match
        found_match = False
        for entry in report_entries:
            if entry.email == email and entry.course_title == course_title:
                entry.course_evaluation = course_evaluation
                found_match = True
        
        # Raise error if no matching entry was found
        if not found_match:
            raise MissingSubmissionData(email, course_title)
        
    del evaluation_df  # Free memory

    # Now we need to compare the report entries to the reference file to find the course ID and hours for each ent
    for entry in report_entries:       
        # Iterate through each sheet in the reference file to find a matching course title and evaluation
        found_match = False

        excel_file = pd.ExcelFile(reference_file)
        sheet_name = None
        for available_sheet in excel_file.sheet_names:
            if isinstance(available_sheet, str) and available_sheet.startswith(entry.state):
                sheet_name = available_sheet
                break
        
        if sheet_name is None:
            raise ReferenceFileMissingSheet(entry.state)

        try:
            sheet_df = pd.read_excel(reference_file, sheet_name=sheet_name, skiprows=2)
            sheet_df["Course Title"]
        except Exception:
            sheet_df = pd.read_excel(reference_file, sheet_name=sheet_name, skiprows=3)

        # Find the last row where Course Title matches the course's title
        matches = sheet_df[sheet_df["Course Title"].str.strip() == entry.course_title.strip()]

        if not matches.empty:
            last_match = matches.iloc[-1]
            entry.course_id = last_match.get("Course Number")
            entry.course_hours = last_match.get("Credit Received")

            expiration_date = last_match.get("Expires")
            if expiration_date is None or pd.isna(expiration_date):
                entry.course_expired = False
            else:
                if expiration_date < entry.course_completed_date:
                    entry.course_expired = True
        else:
            raise ReferenceFileMissingCourse(entry.course_title, sheet_name)
        
    # At this point, all report entries should have course ID and hours populated. We can now generate the report.
    _export_to_excel(report_entries)
    
    print(f"Report generation complete! Generated {len(report_entries)} entries.")
